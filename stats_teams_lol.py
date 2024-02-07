import psycopg2
from sqlalchemy import create_engine
from bs4 import BeautifulSoup
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
from openpyxl import load_workbook
from openpyxl.styles.fills import PatternFill
from openpyxl.styles import Font
from datetime import datetime
import math


data_do_dia = datetime.today().strftime('%d-%m-%Y')

local_db = psycopg2.connect(
    user="postgres",
    password="Santos010802.",
    host="localhost",
    port="5432",
    database="postgres")
engine_local = create_engine('postgresql://postgres:Santos010802.@localhost:5432')
# print('\nConexão bem sucedida!')

# Função para converter o tempo em minutos
def converter_para_minutos(tempo):
    minutos, segundos = map(int, tempo.split(':'))
    return minutos + segundos


def verificar_tabela_se_existe(nome_tabela):
    try:
        query = pd.read_sql_table(f'tbl_stats_teams_lol', engine_local)
        return True
    except:
        return False


def formatar_visual_planilha(link_planilha):
    """Formata as planilhas em EXCEL para melhor visualização do usuário final. Edita tamanho colunas"""
    diretorio_planilha = link_planilha

    # Editar tamanho colunas
    try:
        planilha = load_workbook(diretorio_planilha)
        ws = planilha.active
        # Dimensoes da coluna (width) padrão
        lista_nome_colunas = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R',
                              'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH',
                              'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'A0', 'AP', 'AQ']
        for coluna in lista_nome_colunas:
            ws.column_dimensions[f'{coluna}'].width = 15
            # Colunas especiais
        ws.column_dimensions['e'].width = 25  # 'Media de kills'
        ws.column_dimensions['g'].width = 25  # 'Media de kills'
        ws.column_dimensions['i'].width = 25  # 'Media de kills'

        # Fonte
        fonte = Font(name='Arial', bold=True)
        a1 = ws['A1']
        b1 = ws['B1']
        c1 = ws['C1']
        d1 = ws['D1']
        e1 = ws['E1']
        f1 = ws['F1']
        g1 = ws['G1']
        h1 = ws['H1']
        i1 = ws['I1']
        j1 = ws['J1']
        k1 = ws['K1']
        l1 = ws['L1']
        m1 = ws['M1']
        n1 = ws['N1']
        o1 = ws['O1']
        p1 = ws['P1']
        q1 = ws['Q1']
        r1 = ws['R1']
        s1 = ws['S1']
        t1 = ws['T1']
        u1 = ws['U1']
        v1 = ws['V1']
        w1 = ws['W1']
        x1 = ws['X1']
        y1 = ws['Y1']
        aa1 = ws['AA1']
        ab1 = ws['AB1']
        ac1 = ws['AC1']
        ad1 = ws['AD1']
        ae1 = ws['AE1']
        af1 = ws['AF1']
        ag1 = ws['AG1']
        ah1 = ws['AH1']
        ai1 = ws['AI1']
        aj1 = ws['AJ1']
        ak1 = ws['AK1']
        al1 = ws['AL1']
        am1 = ws['AM1']
        an1 = ws['AN1']
        ao1 = ws['AO1']
        ap1 = ws['AP1']
        aq1 = ws['AQ1']


        # Cor e Fonte da células
        listaa = [a1, b1, c1, d1, e1, f1, g1, h1, i1, j1, k1, l1, m1, n1, o1, p1, q1, r1, s1, t1, u1, v1, w1, x1, y1,
                  aa1, ab1, ac1, ad1, ae1, af1, ag1, ah1, ai1, aj1, ak1, al1, am1, an1, ao1, ap1, aq1]
        for elemento in listaa:
            elemento.fill = PatternFill(start_color="00F1FAEE", fill_type="solid")
            elemento.font = fonte

        # Salva e substitui planilha com alterações feitas no visual das colunas e células
        planilha.save(diretorio_planilha)
    except Exception as e:
        print('[AVISO] Não foi possível formatar o visual da planilha. Verifique se o arquivo')


# ler a base com os times - Id deles
base = pd.read_excel("C:\\Users\\Joao Luiz\\Downloads\\Times.xlsx")
Contagem = 1
lista = []
print(base)


### Verifica se já existe alguma para excluir no DB e fazer uma nova ###
checa_se_tbl_existe = verificar_tabela_se_existe('tbl_stats_teams_lol')
if checa_se_tbl_existe:
    cursor = local_db.cursor()
    sleep(1.5)
    cursor.execute('DROP TABLE tbl_stats_teams_lol')
    local_db.commit()
    print('\nTabela antiga excluida!')
else:
    pass




ie = webdriver.Chrome(executable_path=r"C:\\Users\\Joao Luiz\\Downloads\\ChromeDriver.exe")
print('\nNavegador Iniciado')
# entrar no site
ie.get("https://gol.gg/esports/home/")
sleep(1)

for index, id in base.iterrows():

    codigo_time = str(id['codigo_do_time'])
    time = str(id['time'])
    ligas = str(id['camp'])

    xpath = 'https://gol.gg/teams/team-matchlist/' + str(codigo_time) + '/split-Spring/tournament-ALL/'
    sleep(2.5)

    ie.get(xpath)
    sleep(3)

    # Leitura da tabela
    tabela = ie.find_element(By.XPATH, "/html/body/div/main/div[2]/div/div[3]/div/div/div/table")
    html_content = tabela.get_attribute("outerHTML")
    bs = BeautifulSoup(html_content, 'html.parser')
    table = bs.find(name="table")
    tabela_lida = pd.read_html(str(table), flavor='html5lib')[0]
    tabela_lida['Duration'] = tabela_lida['Duration'].str.replace(':', '.')
    #print(f'\n{tabela_lida["Duration"]}')
    #path = rf'C:\Users\Joao Luiz\Downloads\teste_{data_do_dia}.xlsx'
    #tabela_lida.to_excel(path, sheet_name='Dados')

    # Calcular a média da coluna "duration"
    tabela_lida['Duration'] = tabela_lida['Duration'].astype(float)
    media_duracao = round(tabela_lida['Duration'].mean(), 3)
    #print(f'\nMédia de tempo : {media_duracao}')
    # Ajustar a média se os segundos excederem 60
    pegar_os_minutos = str(media_duracao).split('.')[0]
    pegar_os_segundos = str(media_duracao).split('.')[1]
    if len(pegar_os_segundos) >= 3:
        segundos_tratado = int(pegar_os_segundos) / 10
    #print(f'Os minutos são: {pegar_os_minutos}')
    #print(f'Os segundos são: {segundos_tratado}')
    if segundos_tratado > 60:
        segundos_oficial_final = segundos_tratado - 60
        try:
            segundos_oficial_final = str(segundos_oficial_final).split('.')[0]
        except:
            pass
        if len(segundos_oficial_final) == 1:
            segundos_oficial_final = str(0)+str(segundos_oficial_final)
        #print(f'Segundos oficial final: {segundos_oficial_final}')
        media_duracao_formatada = str(math.ceil(media_duracao))+":"+str(segundos_oficial_final)
    else:
        media_duracao_str = str(media_duracao)
        media_duracao_ajustada = media_duracao_str.replace(".", ":")
        media_duracao_formatada = media_duracao_ajustada
    #print(f'Média de tempo formatada : {media_duracao_formatada}')



    tabela_lida['Total_kills'] = tabela_lida["Unnamed: 2"] + tabela_lida["Unnamed: 7"]
    tabela_lida['Total_torres'] = tabela_lida["Unnamed: 4"] + tabela_lida["Unnamed: 9"]
    tabela_lida['Total_drakes'] = tabela_lida["Unnamed: 5"] + tabela_lida["Unnamed: 10"]

    media_drags = tabela_lida['Total_drakes'].mean()
    media_kills = tabela_lida['Total_kills'].mean()
    media_torres = tabela_lida['Total_torres'].mean()
    ## Kills
    media_kills_time_individual = tabela_lida["Unnamed: 2"].mean()
    media_kills_sofrida_pelo_time = tabela_lida["Unnamed: 7"].mean()
    ## Torres
    media_torres_time_individual = tabela_lida['Unnamed: 4'].mean()
    media_torres_sofrida_pelo_time = tabela_lida['Unnamed: 9'].mean()
    ## Drakes
    media_drags_time_individual = tabela_lida['Unnamed: 5'].mean()
    media_drags_sofrido_pelo_time = tabela_lida['Unnamed: 10'].mean()
    rows = tabela_lida.shape[0]

    # Tempo
    under_27_30 = (tabela_lida.query('Duration < 27.30').shape[0] / rows) * 100
    under_28_30 = (tabela_lida.query('Duration < 28.30').shape[0] / rows) * 100
    under_29_30 = (tabela_lida.query('Duration < 29.30').shape[0] / rows) * 100
    under_30_30 = (tabela_lida.query('Duration < 30.30').shape[0] / rows) * 100
    under_31_30 = (tabela_lida.query('Duration < 31.30').shape[0] / rows) * 100
    under_32_30 = (tabela_lida.query('Duration < 32.30').shape[0] / rows) * 100
    under_33_30 = (tabela_lida.query('Duration < 33.30').shape[0] / rows) * 100
    under_34_30 = (tabela_lida.query('Duration < 34.30').shape[0] / rows) * 100
    #print(f'\n{under_29_30}')

    # drakes
    under_3_5 = (tabela_lida.query('Total_drakes <= 3').shape[0] / rows) * 100
    over_3_5 = (tabela_lida.query('Total_drakes > 3').shape[0] / rows) * 100
    under_4_5 = (tabela_lida.query('Total_drakes <= 4').shape[0] / rows) * 100
    over_4_5 = (tabela_lida.query('Total_drakes > 4').shape[0] / rows) * 100
    over_5_5 = (tabela_lida.query('Total_drakes > 5').shape[0] / rows) * 100

    # torre
    over_12_5 = (tabela_lida.query('Total_torres > 12').shape[0] / rows) * 100
    under_12_5 = (tabela_lida.query('Total_torres <= 12').shape[0] / rows) * 100
    over_11_5 = (tabela_lida.query('Total_torres > 11').shape[0] / rows) * 100
    under_11_5 = (tabela_lida.query('Total_torres <= 11').shape[0] / rows) * 100

    # kills
    over_15_5 = (tabela_lida.query('Total_kills > 15').shape[0] / rows) * 100
    over_16_5 = (tabela_lida.query('Total_kills > 16').shape[0] / rows) * 100
    over_17_5 = (tabela_lida.query('Total_kills > 17').shape[0] / rows) * 100
    over_18_5 = (tabela_lida.query('Total_kills > 18').shape[0] / rows) * 100
    over_19_5 = (tabela_lida.query('Total_kills > 19').shape[0] / rows) * 100
    over_20_5 = (tabela_lida.query('Total_kills > 20').shape[0] / rows) * 100
    over_21_5 = (tabela_lida.query('Total_kills > 21').shape[0] / rows) * 100
    over_22_5 = (tabela_lida.query('Total_kills > 22').shape[0] / rows) * 100
    over_23_5 = (tabela_lida.query('Total_kills > 23').shape[0] / rows) * 100
    over_24_5 = (tabela_lida.query('Total_kills > 24').shape[0] / rows) * 100
    over_25_5 = (tabela_lida.query('Total_kills > 25').shape[0] / rows) * 100
    over_26_5 = (tabela_lida.query('Total_kills > 26').shape[0] / rows) * 100
    over_27_5 = (tabela_lida.query('Total_kills > 27').shape[0] / rows) * 100
    over_28_5 = (tabela_lida.query('Total_kills > 28').shape[0] / rows) * 100
    over_29_5 = (tabela_lida.query('Total_kills > 29').shape[0] / rows) * 100
    over_30_5 = (tabela_lida.query('Total_kills > 30').shape[0] / rows) * 100
    over_31_5 = (tabela_lida.query('Total_kills > 31').shape[0] / rows) * 100
    over_32_5 = (tabela_lida.query('Total_kills > 32').shape[0] / rows) * 100


    porcentagem = "%"
    variavel = "minutos"
    d = {
        'Time': time,
        'Liga': ligas,
        'Jogos': rows,
        'Avg Duração': media_duracao_formatada,
        '% Under 27:30': str(round(under_27_30)) + porcentagem,
        '% Under 28:30': str(round(under_28_30)) + porcentagem,
        '% Under 29:30': str(round(under_29_30)) + porcentagem,
        '% Under 30:30': str(round(under_30_30)) + porcentagem,
        '% Under 31:30': str(round(under_31_30)) + porcentagem,
        '% Under 32:30': str(round(under_32_30)) + porcentagem,
        '% Under 33:30': str(round(under_33_30)) + porcentagem,
        '% Under 34:30': str(round(under_33_30)) + porcentagem,
        'Avg Kills Indiv': round(media_kills_time_individual, 1),
        'Avg Kills Sofridas': round(media_kills_sofrida_pelo_time, 1),
        'Avg Kills': round(media_kills, 2),
        'Avg Towers Indiv': round(media_torres_time_individual, 1),
        'Avg Towers Sofridas': round(media_torres_sofrida_pelo_time, 1),
        'Avg towers': round(media_torres, 1),
        'Avg drakes Indiv': round(media_drags_time_individual, 1),
        'Avg drakes sofridos': round(media_drags_sofrido_pelo_time, 1),
        'Avg drakes': round(media_drags, 1),
        '% Over 11,5 torres': str(round(over_11_5)) + porcentagem,
        '% Under 12,5 torres': str(round(under_12_5)) + porcentagem,
        '% Over 12,5 torres': str(round(over_12_5)) + porcentagem,
        '% Under 3,5 drakes': str(round(under_3_5)) + porcentagem,
        '% Over 3,5 drakes': str(round(over_3_5)) + porcentagem,
        '% Under 4,5 drakes': str(round(under_4_5)) + porcentagem,
        '% Over 4,5 drakes': str(round(over_4_5)) + porcentagem,
        '% Over 5,5 drakes': str(round(over_5_5)) + porcentagem,
        '% Over 18,5': str(round(over_20_5)) + porcentagem,
        '% Over 19,5': str(round(over_20_5)) + porcentagem,
        '% Over 20,5': str(round(over_20_5)) + porcentagem,
        '% Over 21,5': str(round(over_21_5)) + porcentagem,
        '% Over 22,5': str(round(over_22_5)) + porcentagem,
        '% Over 23,5': str(round(over_23_5)) + porcentagem,
        '% Over 24,5': str(round(over_24_5)) + porcentagem,
        '% Over 25,5': str(round(over_25_5)) + porcentagem,
        '% Over 26,5': str(round(over_26_5)) + porcentagem,
        '% Over 27,5': str(round(over_27_5)) + porcentagem,
        '% Over 28,5': str(round(over_28_5)) + porcentagem,
        '% Over 29,5': str(round(over_29_5)) + porcentagem,
        '% Over 30,5': str(round(over_30_5)) + porcentagem,
        '% Over 31,5': str(round(over_31_5)) + porcentagem,
        '% Over 32,5': str(round(over_32_5)) + porcentagem,

    }

    controle = pd.DataFrame(d, index=[Contagem])
    controle.to_sql(f'tbl_stats_teams_lol', engine_local, if_exists='append')
    Contagem += 1

ie.close()
sleep(1.5)

tabela_lol = pd.read_sql_table(f'tbl_stats_teams_lol', engine_local)
sleep(3)

table_lol = pd.DataFrame(tabela_lol)
table_lol.drop(['index'], axis=1, inplace=True)

path = r'C:\Users\Joao Luiz\Downloads\tbl_stats_teams_lol.xlsx'
table_lol.to_excel(path, sheet_name='Dados')

#formatar_visual_planilha(f'C:/Users/Joao Luiz/Downloads/tbl_stats_teams_lol.xlsx')
print('\nTabela salva em excel!')
